######################################################################################################################################################################
# CHANGE LOG                                                                                                                                                         #
# 1. Fixed 'Count of searched words produced each year' bug as new year wasn't added to database since v1.3 started it                                               #
# 2. Add Excel functions to it. Can save in .xls or .xlsx format depending on Excel version 2003 or older or newer													 #
# 3. Include extract data from all text file from folder 																											 #
# 4. Use of tkinter to create a GUI open folder dialog/window for users to easily open/save files or folders														 #
######################################################################################################################################################################

import os
from xlrd import open_workbook
import xlwt
from xlutils.copy import copy
import openpyxl
import tkinter
from tkinter import filedialog

#get only the chat
def clean(dataList):
	while 1:
		if(dataList[0].strip() != '*'):
			del dataList[0]
		else:
			return dataList


#check if the username or year is already a key in the RAM database
def databaseLookup(database, username, year, matchingList):
	#need to create username or year in RAM database
	createUsername = 1
	createYear = 1
	createYearPriKey = 1

	#check if RAM database is empty
	if(len(database) == 0):
		database = {username:{year:{'num_of_words':0}}}
		database[username][year]['turn'] = 0
		database[username][year]['token'] = {}
		database[year] = {}

		for word in matchingList:
			database[username][year]['token'][word] = 0
			database[year][word] = 0

		return database

	#this line onwards are for initializing by usernames
	for key in database:
		if(key == username):
			#don't need to create username in RAM database
			createUsername = 0
			break

	#if need to create a new username, initialize data to 0
	if(createUsername == 1):
		#need to create dictionary layer by layer
		database[username] = {}
		database[username][year] = {}
		database[username][year]['num_of_words'] = 0
		database[username][year]['turn'] = 0
		database[username][year]['token'] = {}

		for word in matchingList:
			database[username][year]['token'][word] = 0

		return database

	for key in database[username]:
		if(key == year):
			#don't need to create username in RAM database
			createYear = 0
			break

	#if need to create a new year, initialize data to 0
	if(createYear == 1):
		#need to create dictionary layer by layer
		database[username][year] = {}
		database[username][year]['num_of_words'] = 0
		database[username][year]['turn'] = 0
		database[username][year]['token'] = {}

		for word in matchingList:
			database[username][year]['token'][word] = 0

	#for 'Count of searched words produced each year'
	for key in database:
		if(key.isdigit()):
			if(key == year):
				createYearPriKey = 0

	#if need to create a new year primary key initialize data to 0
	if(createYearPriKey == 1):
		database[year] = {}
		for i in range(len(matchingList)):
			database[year][matchingList[i]] = 0

	return database


#add data into RAM database
def dataProcessing(database, username, year, chatMessage, matchingList, turn):
	#remove puntuations and emojis, and make them all lowercase
	chatMessage = formatData(chatMessage.lower())
	chatMessageList = chatMessage.split()

	#remove weird characters from username if have
	username = formatUsername(username)

	#check if the username or year is already a key in the RAM database
	database = databaseLookup(database, username, year, matchingList)
	
	#A count of all the words produced by each individual speaker for each year that they sent a message
	#A count of all the “turns” produced by each individual speaker for each year that they sent a message. (I.e. the number of lines/chat bubbles)
	database[username][year]['num_of_words'] += len(chatMessageList)
	database[username][year]['turn'] += turn

	#A count of words(inside the matchingList) produced each year.
	for word in matchingList:
		if(chatMessage.find(word) != -1):
			database[username][year]['token'][word] += 1
			database[year][word] += 1

	return database


#data extraction
def extraction(dataList, namePos, datePos, dateType, yearLen, chatPos, matchingList, database):
	#keep track of line read
	line = 1
	#store last used username
	username = ""
	#keep track of the character that indicater the start of chat bubble
	lineList = dataList[line].split()
	startChar = lineList[chatPos-2][-1]
	#store last used year
	year = ""


	while(dataList[line].strip() != '*'):
		#format data of each line in the text file
		lineList = dataList[line].split()
		#print("line = " + str(line))	#debug

		try:
			#check if it is a new chat bubble or next line of the same chat bubble
			if(lineList[chatPos-2][-1] == startChar):
				#log the username and year incase new line of textfile is under the same chat bubble
				username = formatData(lineList[namePos-1]).strip()
				lineList[datePos-1] = formatData(lineList[datePos-1]).strip()
				year = formatDate(lineList[datePos-1], dateType, yearLen)

				#add data into RAM database
				database = dataProcessing(database, username, year, ' '.join(lineList[chatPos-1:]), matchingList, 1)
			else:
				database = dataProcessing(database, username, year, ' '.join(lineList), matchingList, 0)
		except IndexError:
			database = dataProcessing(database, username, year, ' '.join(lineList), matchingList, 0)
		except Exception as e:
			pass

		#to read next line of data
		line += 1
		#break

	return database


#clear out puntuations and emojis
def formatData(chatMessage):
	for i in range(len(chatMessage)):
		#maximize efficiency through shortcircuiting with characters with higher possibility of appearing
		if((ord(chatMessage[i]) > 96 and ord(chatMessage[i]) < 123) or (ord(chatMessage[i]) > 64 and ord(chatMessage[i]) < 91) or \
			(ord(chatMessage[i]) > 47 and ord(chatMessage[i]) < 58) or ord(chatMessage[i]) == 39 or ord(chatMessage[i]) == 47):
			continue
		else:
			chatMessage = chatMessage.replace(chatMessage[i], ' ')
	

	#get list of each line
	return chatMessage


#extracting year depending on date format
def formatDate(date, dateType, yearLen):
	if(dateType == 1):
		#extract from the back have to be negative
		yearLen *= -1
		return date[yearLen:] if len(date[yearLen:]) == 4 else "20" + date[yearLen:]
	else:
		return date[:yearLen] if len(date[:yearLen]) == 4 else "20" + date[:yearLen]


#get chat format
def formatInfo():
	while 1:
		namePos = input("Please input the position of the name (eg. A/F/CH/18): ")

		if(RepresentsInt(namePos) == True):
			namePos = int(namePos)
			break
		else:
			print("Please input numbers only.\n")


	while 1:
		datePos = input("Please input the position of the date (eg. 11/05/18): ")

		if(RepresentsInt(datePos) == True):
			datePos = int(datePos)
			break
		else:
			print("Please input numbers only.\n")


	while 1:
		print("\n1. Day/Month/Year or Month/Day/Year")
		print("2. Year/Day/Month or Year/Month/Day")
		dateType = input("Please input the format of the date: ")

		if(RepresentsInt(dateType) == True):
			dateType = int(dateType)
			if(dateType != 1 and dateType != 2):
				print("Please the following options only.\n")
				continue
			break
		else:
			print("Please input numbers only.\n")


	while 1:
		yearLen = input("\nPlease input the number of digits that represents the year (eg. 2018 is 4 digits, 18 is 2 digits): ")

		if(RepresentsInt(yearLen) == True):
			yearLen = int(yearLen)
			break
		else:
			print("Please input numbers only.\n")


	while 1:
		chatPos = input("Please input the position of the start of the message: ")

		if(RepresentsInt(chatPos) == True):
			chatPos = int(chatPos)
			break
		else:
			print("Please input numbers only.\n")


	return namePos, datePos, dateType, yearLen, chatPos

#clear out weird unicode characters from username
def formatUsername(username):
	for i in range(len(username)):
		if not(ord(username[i]) > 31 and ord(username[i]) < 127):
			username = username.replace(username[i], 'x')

	return username


#get location of text file
def getAFile():
	while 1:
		#clear screen (cross platform)
		os.system('cls' if os.name == 'nt' else 'clear')
		print("To exit the program, press CTRL+C on your keyboard.")
		print("\nPlease input the location of the text file: ", end="")

		#create tkinter object
		tk = tkinter.Tk()
		#open save file dialog GUI to get filename to save
		fileLoc = filedialog.askopenfilename(defaultextension = '.txt', filetype=[("Text document", "*.txt")])
		#close the mini annoying tk window that will appear too when opening 'save file dialog'
		tk.destroy()
		#remove full path of file and get only filename
		if(fileLoc is ""):
			continue

		#display filename
		print(fileLoc)

		return fileLoc


#get location of folder then get all the next files
def getAFolder():
	while 1:
		#clear screen (cross platform)
		os.system('cls' if os.name == 'nt' else 'clear')
		print("To exit the program, press CTRL+C on your keyboard.")
		print("\nPlease input the location of the folder: ", end="")

		#create tkinter object
		tk = tkinter.Tk()
		#open save file dialog GUI to get filename to save
		folderLoc = filedialog.askdirectory()
		#close the mini annoying tk window that will appear too when opening 'save file dialog'
		tk.destroy()
		#remove full path of file and get only filename
		if(folderLoc == ""):
			continue
		
		#display filename
		print(folderLoc)

		fileLoc = os.listdir(folderLoc)

		#remove files that are not text file
		for i in range(len(fileLoc)):
			if(fileLoc[i][-4:] != ".txt"):
				fileLoc.remove(file)
			else:
				fileLoc[i] = folderLoc + "/" + fileLoc[i]
		
		return fileLoc


#get words to match
def getWordList(matchingList):
	word = ""

	#clear screen (cross platform)
	os.system('cls' if os.name == 'nt' else 'clear')
	print("Below are the list of words to search for.")
	print("To delete the word, input \"del <number>\". Example: \"del 2\" to delete the 2nd word. ")
	print("To end, input \"q\"\n")

	#print all the current list choosen
	for i in range(len(matchingList)):
		print(str(i+1) + ". " + matchingList[i])

	word = input("Words to add in: ").lower()


	#check if user wants to quit
	while(word != 'q'):
		try:
			if(word[:4] == 'del '):
				del matchingList[int(word[4:])-1]
			else:
				#add word to matchingList
				matchingList.append(word)
		except IndexError:
			#add word to matchingList
			matchingList.append(word)
		except ValueError:
			#add word to matchingList
			matchingList.append(word)
		except Exception as e:
			print(e)
			print("Screenshot and email the issue to haojie_34@hotmail.com\n")

		#clear screen (cross platform)
		os.system('cls' if os.name == 'nt' else 'clear')
		print("Below are the list of words to search for.")
		print("To delete the word, input \"del <number>\". Example: \"del 2\" to delete the 2nd word. ")
		print("To quit, input \"q\"\n")

		#print all the current list choosen
		for i in range(len(matchingList)):
			print(str(i+1) + ". " + matchingList[i])

		word = input("Words to add in: ").lower()

	#clear screen (cross platform)
	os.system('cls' if os.name == 'nt' else 'clear')

	return matchingList


#display result and save result to text file
def outputResults(database, totalSum, matchingList, fileLoc):
	#to display results on CLI and also add into text file in the same layout
	resultList = []
	
	#printing results
	for key, value in database.items():
		#if true, it means content is by username
		if(not(RepresentsInt(key))):
			#print username
			resultList.append("[" + key + "]\n")
			for k, val in value.items():
				#print year
				resultList.append("year = " + k + "\n")
				for k1, val1 in val.items():
					#print either no. of words or number of turns and check if it's tokens (internal dictionary)
					if(k1 == 'num_of_words'):
						resultList.append("Number of words in that year = " + str(val1) + "\n")
					elif(k1 == 'turn'):
						resultList.append("Number of \"turns\" in that year = "  + str(val1) + "\n")
					else:
						for k2, val2 in val1.items():
							resultList.append("Number of \'"+ k2 +"\' in that year = "  + str(val2) + "\n")

			resultList.append("\n\n")


	#extra spacing to seperate number of matched words
	resultList.append("\n[Count of searched words produced each year]\n")

	for key, value in database.items():
		#if true, it means content is by year
		if(RepresentsInt(key)):
			#print year
			resultList.append("[" + key + "]\n")
			#incase matchingList is empty, there will be no data in dict
			if(len(value) != 0):
				for k, val in value.items():
					#print word
					resultList.append(k + " = " + str(val) + "\n")
			else:
				resultList.append("NIL\n")
			resultList.append("\n\n")


	#extra spacing to seperate total data
	resultList.append("\n[Total count per speaker]\n")

	for name in totalSum:
		resultList.append("[" + name + "]\n")
		resultList.append("Total words = " + str(totalSum[name]['num_of_words']) + "\n")
		resultList.append("Total \"turns\" = " + str(totalSum[name]['turn']) + "\n")
		#to get the tokens cause varies
		for token, value in totalSum[name]['token'].items():
			resultList.append("Total \'" + token + "\' = " + str(totalSum[name]['token'][token]) + "\n")
		resultList.append("\n\n")

	for i in range(len(resultList)):
		print(resultList[i], end="")


	save = input("Do you need to store the results a text/excel file? (y/n)\n").lower() 

	#clear screen (cross platform)
	os.system('cls' if os.name == 'nt' else 'clear')

	if(save[0] == 'y'):
		writing(database, totalSum, matchingList, resultList, fileLoc, "")


def reading(fileLoc):
	try:
		#whatsapp chat txt file is in UTF-8 with BOM
		myFile = open(fileLoc, "r", encoding="utf-8-sig")

	except UnicodeDecodeError:
		myFile = open(fileLoc, "r")
	except FileNotFoundError:
		print("File Not Found Error")
	except Exception as e:
		print(e)
		print("Screenshot and email the issue to haojie_34@hotmail.com\n")

	else:
		dataList = myFile.readlines()
		myFile.close()
		return dataList


#check if is integer
def RepresentsInt(s):
    try: 
        int(s)
        return True
    except ValueError:
        return False


#Additional data processing
def totalUpData(database, matchingList):
	#contains total words per person, total "turns" per person without 'age incrementing every year' issue
	totalSum = {}

	for key, value in database.items():
		#true if not 'year'
		if(not(RepresentsInt(key))):
			#get the names without the age
			nameWithoutAge = key.split('/')
			nameWithoutAge = '/'.join(nameWithoutAge[:-1])
			#create new key and layer if name (key) does not exist and initialize to 0
			if nameWithoutAge not in totalSum:
				totalSum[nameWithoutAge] = {'num_of_words': 0, 'turn': 0, 'token': {}}
				for word in matchingList:
					totalSum[nameWithoutAge]['token'][word] = 0
			#k will be year in name's dictionary
			for k, v in value.items():
				#k1 will be 'num_of_words', 'turn' and 'token'
				for k1, v1 in v.items():
					#other v1 maybe values of 'num_of_words' and 'turn'
					if(k1 != 'token'):
						#adding 'num_of_words' and 'turn' counts into totalSum
						totalSum[nameWithoutAge][k1] += v[k1]
					else:
						for k2 in v1:
							#adding token counts into totalSum when reach final layer
							totalSum[nameWithoutAge]['token'][k2] += v1[k2]

	return database, totalSum


#write result to old excel file format
#https://hackernoon.com/working-with-spreadsheets-using-python-part-1-380a120387f
def writing(database, totalSum, matchingList, resultList, fileLoc, fileName):
	print("To exit the program, press CTRL+C on your keyboard.\n")
	if(fileName == ""):
		#https://www.youtube.com/watch?v=iUmqLGUktek
		print("Input the name of your file: ", end="")
		#create tkinter object
		tk = tkinter.Tk()
		#open save file dialog GUI to get filename to save
		fileName = filedialog.asksaveasfilename(defaultextension = '.xlsx', 
			filetype=[("Excel Workbook", "*.xlsx"), ("Excel 97-2003 Workbook", "*.xls"), ("Text document", "*.txt")])
		#close the mini annoying tk window that will appear too when opening 'save file dialog'
		tk.destroy()

	#return to main menu if user clicks cancel in save file dialog
	if(fileName == ""):
		return
	print(fileName)


	#clear screen (cross platform)
	os.system('cls' if os.name == 'nt' else 'clear')

	#check file type as newer than year 2003 version (.xlsx) uses openpyxl
	#year 2003 or older versions are .xls and need to ue xlrd, xlutils.copy, and xlwt
	#may return Excel file location
	if(fileName[-5:] == ".xlsx"):
		return writing_excel(fileName, fileLoc, database, totalSum, matchingList, True)
	elif(fileName[-4:] == ".xls"):
		return writing_excel(fileName, fileLoc, database, totalSum, matchingList, False)
	else:
		writing_text(fileName, resultList, fileLoc)


#write result to text file
def writing_text(fileName, resultList, fileLoc):
	print("To exit the program, press CTRL+C on your keyboard.\n")

	#clear content of file to write or create a new file to write
	myFile = open(fileName, "+w")
	myFile.writelines(resultList)
	myFile.close()

	#clear screen (cross platform)
	os.system('cls' if os.name == 'nt' else 'clear')
	print("Result is successfully saved!")


def writing_excel(fileName, fileLoc, database, totalSum, matchingList, is_xlsx):
	#to store the position of the word "END" to know which row to append new content
	#.xlsx starts from 1 while .xls starts from 0
	END_Pos = 1

	try:
		#.xlsx uses openpyxl while .xls uses xlrd, xlwt, and xlutil
		if(is_xlsx):
			wb = openpyxl.load_workbook(fileName)
		else:
			#to open .xls excel file and load into read buffer using xlrd
			#on_demand=True will use less memory
			rb = open_workbook(fileName, formatting_info=True)

	except FileNotFoundError:
		if(is_xlsx):
			#create new Excel 2010 or later
			wb = openpyxl.Workbook()
			#get the 1st sheet
			sheet = wb.active
			#newer version of Excel is different as start from 1 instead of 0
			sheet.cell(row=1, column=1, value="END")
			wb.save(fileName)
			wb = openpyxl.load_workbook(fileName)
		else:
			#if not found, create that file then open using rb to make the code after EXCEPT block reuseable
			createExcelFile = xlwt.Workbook()
			sheet = createExcelFile.add_sheet("Sheet1")
			#write to row 0, column 0 the word "END"
			sheet.write(0, 0, "END")
			createExcelFile.save(fileName)
			rb = open_workbook(fileName, formatting_info=True)

	if(is_xlsx):
		#get the 1st sheet (tab)
		w_sheet = wb.active
	else:
		#get the data from the 1st sheet (tab) since index we set to 0
		r_sheet = rb.sheet_by_index(0)
		#copy read buffer content and create a writable buffer
		#another reason to copy as xlwt is for .xls which cannot append to file
		#hence copy the original content to buffer(RAM), append from there then overwrite the .xls with the content
		wb = copy(rb)
		w_sheet = wb.get_sheet(0)

	while(1):
		if(is_xlsx):
			if(w_sheet.cell(row=END_Pos, column=1).value == "END"):
				break
		else:
			#find the position of the word "END"
			if(r_sheet.cell(END_Pos-1, 0).value == "END"):
				#due to .xlsx starts from 1 instead of 0 (for .xls), we need to -1 for .xls
				END_Pos -= 1
				break
		
		END_Pos += 1


	#to transverse within the excel file to write content
	excelRowPointer = END_Pos
	excelColumnPointer = 2 if(is_xlsx) else 1

	#write filename
	#remove file path and get only the name
	fileLoc = fileLoc.split('/')
	fileLoc = fileLoc[-1]
	if(is_xlsx):
		w_sheet.cell(row=END_Pos, column=1, value=fileLoc[:-4])
		w_sheet.cell(row=excelRowPointer, column=excelColumnPointer, value="Name")
	else:
		w_sheet.write(END_Pos, 0, fileLoc[:-4])
		w_sheet.write(excelRowPointer, excelColumnPointer, "Name")

	################	for headers		###############
	#to write the year headers
	excelColumnPointer += 1
	#know the number of times to write the same year
	excelNumOfYearsHeader = 2 + len(matchingList)
	#write years into cell from database variable
	for key in database:
		if(key.isdigit()):
			for i in range(excelNumOfYearsHeader):
				#write name (vertically)
				if(is_xlsx):
					w_sheet.cell(row=excelRowPointer, column=excelColumnPointer, value=key)
				else:
					w_sheet.write(excelRowPointer, excelColumnPointer, key)
				#the rest below writes header (horizontally)
				if(i == 0):
					if(is_xlsx):
						w_sheet.cell(row=excelRowPointer+1, column=excelColumnPointer, value="Words")
					else:
						w_sheet.write(excelRowPointer+1, excelColumnPointer, "Words")
				elif(i == 1):
					if(is_xlsx):
						w_sheet.cell(row=excelRowPointer+1, column=excelColumnPointer, value="TURNS")
					else:
						w_sheet.write(excelRowPointer+1, excelColumnPointer, "TURNS")
				else:
					#write the tokens
					if(is_xlsx):
						w_sheet.cell(row=excelRowPointer+1, column=excelColumnPointer, value= matchingList[i-2])
					else:
						w_sheet.write(excelRowPointer+1, excelColumnPointer, matchingList[i-2])

				#+1 to pointer to move to next column
				excelColumnPointer += 1


	#pointer for later use for data of Total words, Total TURNS, etc
	totalDataColumnPointer = excelColumnPointer
	#create headers for Total
	if(is_xlsx):
		w_sheet.cell(row=excelRowPointer+1, column=excelColumnPointer, value="Total words")
		w_sheet.cell(row=excelRowPointer+1, column=excelColumnPointer+1, value="Total TURNS")
	else:
		w_sheet.write(excelRowPointer+1, excelColumnPointer, "Total words")
		w_sheet.write(excelRowPointer+1, excelColumnPointer+1, "Total TURNS")
	excelColumnPointer += 2

	#write headers for Total TOKENS
	for i in range(len(matchingList)):
		if(is_xlsx):
			w_sheet.cell(row=excelRowPointer+1, column=excelColumnPointer, value="Total \"" + matchingList[i] + "\"")
		else:
			w_sheet.write(excelRowPointer+1, excelColumnPointer, "Total \"" + matchingList[i] + "\"")
		#move to next column
		excelColumnPointer += 1


	#reset position of pointers
	#to transverse within the excel file to write content
	excelRowPointer = END_Pos
	excelColumnPointer = 2 if(is_xlsx) else 1


	################	for content		###############
	#include one cell spacing before writing list of names
	excelRowPointer += 2
	for key, value in database.items():
		#write names into cell from database variable
		if(not(key.isdigit())):
			if(is_xlsx):
				w_sheet.cell(row=excelRowPointer, column=excelColumnPointer, value=key)
			else:
				w_sheet.write(excelRowPointer, excelColumnPointer, key)
			#to write the data of that name such as Num Of Words, etc
			#k will be year of data by that name
			excelColumnPointerTemp = excelColumnPointer + 1
			for k, v in value.items():
				for k1, v1 in v.items():
					if(k1 == 'num_of_words'):
						if(is_xlsx):
							w_sheet.cell(row=excelRowPointer, column=excelColumnPointerTemp, value=v1)
						else:
							w_sheet.write(excelRowPointer, excelColumnPointerTemp, v1)
					elif(k1 == 'turn'):
						if(is_xlsx):
							w_sheet.cell(row=excelRowPointer, column=excelColumnPointerTemp+1, value=v1)
						else:
							w_sheet.write(excelRowPointer, excelColumnPointerTemp+1, v1)
						excelColumnPointerTemp += 2
					else:
						#write data for TOKENS
						for k2, v2 in v1.items():
							if(is_xlsx):
								w_sheet.cell(row=excelRowPointer, column=excelColumnPointerTemp, value=v2)
							else:
								w_sheet.write(excelRowPointer, excelColumnPointerTemp, v2)
							#move temp column pointer to the right by one cell
							excelColumnPointerTemp += 1

			#+1 to pointer to move to next row
			excelRowPointer += 1

	#write names into cell from totalSum variable
	for key, value in totalSum.items():
		if(is_xlsx):
			w_sheet.cell(row=excelRowPointer, column=excelColumnPointer, value=key)
		else:
			w_sheet.write(excelRowPointer, excelColumnPointer, key)
		totalDataColumnPointerTemp = totalDataColumnPointer
		#write content of total data by that name
		for k, v in value.items():
			if(k == 'num_of_words'):
				if(is_xlsx):
					w_sheet.cell(row=excelRowPointer, column=totalDataColumnPointerTemp, value=v)
				else:
					w_sheet.write(excelRowPointer, totalDataColumnPointerTemp, v)
			elif(k == 'turn'):
				if(is_xlsx):
					w_sheet.cell(row=excelRowPointer, column=totalDataColumnPointerTemp+1, value=v)
				else:
					w_sheet.write(excelRowPointer, totalDataColumnPointerTemp+1, v)
				totalDataColumnPointerTemp += 2
			else:
				for k1, v1 in v.items():
					if(is_xlsx):
						w_sheet.cell(row=excelRowPointer, column=totalDataColumnPointerTemp, value=v1)
					else:
						w_sheet.write(excelRowPointer, totalDataColumnPointerTemp, v1)
					#move Total Data content column pointer by one cell to the right
					totalDataColumnPointerTemp += 1

		#+1 to pointer to move to next row
		excelRowPointer += 1


	#for Total's row
	if(is_xlsx):
		w_sheet.cell(row=excelRowPointer, column=excelColumnPointer, value="Total")
	else:
		w_sheet.write(excelRowPointer, excelColumnPointer, "Total")
	excelColumnPointer = 5 if(is_xlsx) else 4
	#for Total's data in the row
	for key, value in database.items():
		if(key.isdigit()):
			#k is the TOKENs' name and v is the total amount of each TOKEN
			for k, v in value.items():
				if(is_xlsx):
					w_sheet.cell(row=excelRowPointer, column=excelColumnPointer, value=v)
				else:
					w_sheet.write(excelRowPointer, excelColumnPointer, v)
				#move column pointer by one cell to the right for next TOKEN's total data
				excelColumnPointer += 1

			excelColumnPointer += 2

	if(is_xlsx):
		w_sheet.cell(row=excelRowPointer+3, column=1, value="END")
	else:
		w_sheet.write(excelRowPointer+3, 0, "END")

	try:
		#save file
		wb.save(fileName)
	except PermissionError:
		print("Please close file before saving")
		print("Press enter to continue...")
		input()
		#save file
		wb.save(fileName)


	#clear screen (cross platform)
	os.system('cls' if os.name == 'nt' else 'clear')

	return fileName


def mainFork(namePos, datePos, dateType, yearLen, chatPos, dataList, matchingList, fileLoc, excelFileLoc, isFromFolder):
	#to store each person's data
	database = {}

	#get only the chat
	dataList = clean(dataList)

	if(len(matchingList) == 0):
		#get chat format
		namePos, datePos, dateType, yearLen, chatPos = formatInfo()

		#get words to match
		matchingList = getWordList(matchingList)

	#data extraction
	database = extraction(dataList, namePos, datePos, dateType, yearLen, chatPos, matchingList, database)

	#addtitional data process
	database, totalSum = totalUpData(database, matchingList)
	
	#if from folder, don't have to display result and store Excel file location for reuse of data text file from the same folder
	if(isFromFolder):
		excelFileLoc = writing(database, totalSum, matchingList, "", fileLoc, excelFileLoc)
	else:
		#display result and save result to text file
		outputResults(database, totalSum, matchingList, fileLoc)

	#return many data for reuse by data text files from the same folder if is extracted from folder
	return namePos, datePos, dateType, yearLen, chatPos, matchingList, excelFileLoc


#introduction
print("Welcome to dataScript.")
print("Please follow the instructions below to generate the desire results.")
print("For any enquiry, please email to hadessoulx@gmail.com\n")


#main program
while 1:
	#main menu
	try:
		print("\n1. Extract data from a file")
		print("2. Extract data from a folder of files")
		print("3. Quit")
		choice = int(input("Input choice: "))

	except Exception:
		#clear screen (cross platform)
		os.system('cls' if os.name == 'nt' else 'clear')
		print("Please input only the choices below!")
		continue

	else:
		if(choice < 1 or choice > 3):
			#clear screen (cross platform)
			os.system('cls' if os.name == 'nt' else 'clear')
			print("Please input only the choices below!")
			continue
		elif(choice == 3):
			break

		#data positions
		namePos = 0 
		datePos = 0 
		dateType = 0
		yearLen = 0
		chatPos = 0
		#store data of words to match
		matchingList = []
		#store the location of Excel file
		excelFileLoc = ""

		#Loading content to buffer
		if(choice == 1):
			fileLoc = getAFile()
		else:
			fileLoc = getAFolder()

		if(isinstance(fileLoc, list)):
			for file in fileLoc:
				dataList = reading(file)
				namePos, datePos, dateType, yearLen, chatPos, matchingList, excelFileLoc = mainFork(namePos, 
					datePos, dateType, yearLen, chatPos, dataList, matchingList, file, excelFileLoc, True)
		else:
			dataList = reading(fileLoc)
			mainFork(namePos, datePos, dateType, yearLen, chatPos, dataList, matchingList, fileLoc, excelFileLoc, False)

